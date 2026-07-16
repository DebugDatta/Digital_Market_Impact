import logging
import warnings
from pathlib import Path
from typing import Dict, List, Tuple
import numpy as np
import pandas as pd
import yfinance as yf
from scipy import stats as sp_stats
from statsmodels.tsa.stattools import adfuller
from statsmodels.stats.diagnostic import het_arch
from arch import arch_model
import ruptures as rpt
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.INFO, format='%(asctime)s | %(levelname)s | %(message)s', datefmt='%H:%M:%S')
log = logging.getLogger(__name__)

ALGORITHM_TICKERS = ['NYKAA.NS', 'ETERNAL.NS', 'SWIGGY.NS', 'PAYTM.NS']
TRADITIONAL_TICKERS = ['HINDUNILVR.NS', 'BRITANNIA.NS', 'DABUR.NS', 'MARICO.NS', 'ITC.NS']
ALL_TICKERS = ALGORITHM_TICKERS + TRADITIONAL_TICKERS + ['^NSEI']
BENCHMARK = '^NSEI'
NIFTY_LABEL = 'NIFTY50'
MIN_OBS = 252
ROLLING_WINDOW = 21
TRADING_DAYS = 252
PELT_PENALTY = 5.0
FIG_SIZE = (12, 6)
FIG_DPI = 200
COLORS = plt.cm.tab10
FIGURE_DIR = Path('figures')
EXCEL_PATH = Path('results.xlsx')

def _short(name: str) -> str:
    return name.replace('.NS', '').replace('^NSEI', NIFTY_LABEL)

def download_data(tickers: List[str]) -> pd.DataFrame:
    log.info('Downloading price data for %d tickers ...', len(tickers))
    raw = yf.download(tickers, period='max', progress=False, auto_adjust=False)
    if 'Close' in raw.columns:
        prices = raw['Close'].copy()
    else:
        prices = raw.xs('Close', axis=1, level=0).copy()
    log.info('Raw shape: %s', prices.shape)
    return prices

def align_and_clean(prices: pd.DataFrame) -> pd.DataFrame:
    log.info('Cleaning & aligning data ...')
    prices = prices.ffill().dropna(how='any')
    prices = prices.loc[~prices.index.duplicated(keep='first')]
    log.info('Cleaned shape: %s, date range: %s to %s', prices.shape, prices.index[0].date(), prices.index[-1].date())
    return prices

def compute_returns(prices: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.Series]:
    log.info('Computing returns & volatility ...')
    log_ret = np.log(prices / prices.shift(1)).dropna()
    pct_ret = prices.pct_change().dropna()
    rolling_vol = log_ret.rolling(ROLLING_WINDOW).std() * np.sqrt(TRADING_DAYS) * 100
    ann_vol = log_ret.std() * np.sqrt(TRADING_DAYS) * 100
    log.info('Log returns shape: %s', log_ret.shape)
    return log_ret, pct_ret, rolling_vol, ann_vol

def filter_by_obs(df: pd.DataFrame, label: str) -> pd.DataFrame:
    all_cols = [c for c in df.columns if c != BENCHMARK]
    valid = [c for c in all_cols if df[c].notna().sum() >= MIN_OBS]
    dropped = set(all_cols) - set(valid)
    if dropped:
        log.warning('%s — dropping (obs < %d): %s', label, MIN_OBS, [_short(t) for t in sorted(dropped)])
    keep = valid + [BENCHMARK]
    out = df[keep]
    log.info('%s stocks after filter: %d', label, out.shape[1])
    return out

def calc_descriptive_stats(log_ret: pd.DataFrame, ann_vol: pd.Series) -> pd.DataFrame:
    stats = pd.DataFrame({
        'Mean Return (%)': log_ret.mean() * 100,
        'Std Dev (%)': log_ret.std() * 100,
        'Annualized Vol (%)': ann_vol,
        'Skewness': log_ret.skew(),
        'Kurtosis': log_ret.kurtosis() + 3,
        'Minimum (%)': log_ret.min() * 100,
        'Maximum (%)': log_ret.max() * 100,
    })
    stats.index = [_short(s) for s in stats.index]
    return stats

def run_adf(series: pd.Series) -> Tuple[float, float]:
    res = adfuller(series.dropna(), maxlag=12, autolag='AIC')
    return float(res[0]), float(res[1])

def run_arch_lm(series: pd.Series, nlags: int = 5) -> Tuple[float, float]:
    res = het_arch(series.dropna(), nlags=nlags)
    return float(res[0]), float(res[1])

def run_jb(series: pd.Series) -> Tuple[float, float]:
    stat, pv = sp_stats.jarque_bera(series.dropna())
    return float(stat), float(pv)

def build_stats_table(log_ret: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for col in log_ret.columns:
        s = log_ret[col].dropna()
        adf_stat, adf_p = run_adf(s)
        arch_stat, arch_p = run_arch_lm(s)
        jb_stat, jb_p = run_jb(s)
        rows.append({
            'Stock': _short(col),
            'ADF Statistic': adf_stat,
            'ADF p-value': adf_p,
            'ARCH-LM Statistic': arch_stat,
            'ARCH-LM p-value': arch_p,
            'JB Statistic': jb_stat,
            'JB p-value': jb_p,
        })
    return pd.DataFrame(rows)

def fit_garch11(log_ret: pd.DataFrame) -> Tuple[pd.DataFrame, dict]:
    log.info('Fitting GARCH(1,1) for each stock ...')
    results = []
    models: Dict[str, arch_model.ARCHModelResult] = {}
    for col in log_ret.columns:
        s = log_ret[col].dropna() * 100
        lbl = _short(col)
        try:
            am = arch_model(s, mean='Zero', vol='GARCH', p=1, q=1, dist='normal')
            res = am.fit(disp='off', show_warning=False)
            omega = float(res.params.get('omega', np.nan))
            alpha = float(res.params.get('alpha[1]', np.nan))
            beta = float(res.params.get('beta[1]', np.nan))
            persist = alpha + beta if not (np.isnan(alpha) or np.isnan(beta)) else np.nan
            results.append({
                'Stock': lbl, 'Omega': omega, 'Alpha': alpha,
                'Beta': beta, 'Persistence': persist,
                'AIC': res.aic, 'BIC': res.bic, 'Log-Likelihood': res.loglikelihood,
            })
            models[col] = res
            log.info('  %12s  \u03c9=%.6f  \u03b1=%.4f  \u03b2=%.4f  \u03b1+\u03b2=%.4f', lbl, omega, alpha, beta, persist)
        except Exception as exc:
            log.warning('  %12s  GARCH failed: %s', lbl, exc)
            results.append({
                'Stock': lbl, 'Omega': np.nan, 'Alpha': np.nan,
                'Beta': np.nan, 'Persistence': np.nan,
                'AIC': np.nan, 'BIC': np.nan, 'Log-Likelihood': np.nan,
            })
    return pd.DataFrame(results), models

def group_volatility_test(ann_vol: pd.Series) -> Tuple[float, float, float, float]:
    alg_vals = ann_vol[[t for t in ALGORITHM_TICKERS if t in ann_vol.index]].values
    trad_vals = ann_vol[[t for t in TRADITIONAL_TICKERS if t in ann_vol.index]].values
    t_stat, t_p = sp_stats.ttest_ind(alg_vals, trad_vals, equal_var=False)
    l_stat, l_p = sp_stats.levene(alg_vals, trad_vals)
    return float(t_stat), float(t_p), float(l_stat), float(l_p)

def detect_breaks(log_ret: pd.DataFrame, penalty: float = PELT_PENALTY) -> pd.DataFrame:
    log.info('Detecting structural breaks (PELT, pen=%.1f) ...', penalty)
    max_n = 0
    breaks: Dict[str, List[str]] = {}
    for col in log_ret.columns:
        s = (log_ret[col].dropna().values * 100).reshape(-1, 1)
        algo = rpt.Pelt(model='l2', min_size=15).fit(s)
        bkps = algo.predict(pen=penalty)
        dates = [str(log_ret.index[b - 1].date()) for b in bkps if b < len(log_ret)]
        lbl = _short(col)
        breaks[lbl] = dates
        max_n = max(max_n, len(dates))
        log.info('  %12s: %d breaks', lbl, len(dates))
    cols = [f'Break {i + 1}' for i in range(max_n)]
    data = {}
    for lbl, dates in breaks.items():
        row = dates + [''] * (max_n - len(dates))
        data[lbl] = row
    result = pd.DataFrame(data).T
    result.columns = cols
    result.index.name = 'Stock'
    return result, breaks

def _save_fig(name: str):
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    path = FIGURE_DIR / name
    plt.savefig(str(path), dpi=FIG_DPI, bbox_inches='tight')
    plt.close()
    log.info('Saved figure: %s', path)

def figure_1_prices(prices: pd.DataFrame):
    log.info('Plotting Figure 1: Normalized prices ...')
    norm = prices / prices.iloc[0] * 100
    fig, ax = plt.subplots(figsize=FIG_SIZE)
    for i, col in enumerate(norm.columns):
        ax.plot(norm.index, norm[col], color=COLORS(i / 10), lw=1.2, label=_short(col))
    ax.set_title('Stock Price Comparison (Normalized to 100)', fontsize=14)
    ax.set_ylabel('Normalized Price')
    ax.legend(ncol=2, fontsize=8, framealpha=0.7)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    _save_fig('fig1_price_comparison.png')

def figure_2_rolling_vol(rolling_vol: pd.DataFrame):
    log.info('Plotting Figure 2: Rolling 21-day volatility ...')
    fig, ax = plt.subplots(figsize=FIG_SIZE)
    for i, col in enumerate(rolling_vol.columns):
        ax.plot(rolling_vol.index, rolling_vol[col], color=COLORS(i / 10), lw=0.9, alpha=0.85, label=_short(col))
    ax.set_title('Rolling 21-Day Annualized Volatility (%)', fontsize=14)
    ax.set_ylabel('Annualized Volatility (%)')
    ax.legend(ncol=2, fontsize=8, framealpha=0.7)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    _save_fig('fig2_rolling_volatility.png')

def figure_3_boxplot(log_ret: pd.DataFrame):
    log.info('Plotting Figure 3: Boxplot by group ...')
    alg_cols = [c for c in ALGORITHM_TICKERS if c in log_ret.columns]
    trad_cols = [c for c in TRADITIONAL_TICKERS if c in log_ret.columns]
    alg_data = log_ret[alg_cols].values.flatten() * 100
    trad_data = log_ret[trad_cols].values.flatten() * 100
    alg_data = alg_data[~np.isnan(alg_data)]
    trad_data = trad_data[~np.isnan(trad_data)]
    fig, ax = plt.subplots(figsize=(8, 6))
    bp = ax.boxplot([alg_data, trad_data], labels=['Algorithm-Driven', 'Traditional'], widths=0.5, patch_artist=True, medianprops={'color': 'black', 'lw': 2})
    bp['boxes'][0].set_facecolor('#FF6B6B')
    bp['boxes'][1].set_facecolor('#4ECB71')
    ax.set_title('Daily Returns by Group (%)', fontsize=14)
    ax.set_ylabel('Daily Return (%)')
    ax.grid(True, axis='y', alpha=0.3)
    fig.tight_layout()
    _save_fig('fig3_boxplot_returns.png')

def figure_4_corr_heatmap(log_ret: pd.DataFrame):
    log.info('Plotting Figure 4: Correlation heatmap ...')
    corr = log_ret.corr()
    labels = [_short(c) for c in corr.columns]
    n = len(labels)
    fig, ax = plt.subplots(figsize=(10, 9))
    im = ax.imshow(corr.values, cmap='RdBu_r', vmin=-1, vmax=1, aspect='equal')
    ax.set_xticks(range(n))
    ax.set_yticks(range(n))
    ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
    ax.set_yticklabels(labels, fontsize=8)
    for i in range(n):
        for j in range(n):
            val = corr.values[i, j]
            color = 'white' if abs(val) > 0.5 else 'black'
            ax.text(j, i, f'{val:.2f}', ha='center', va='center', fontsize=7, color=color)
    cbar = fig.colorbar(im, ax=ax, shrink=0.75)
    cbar.set_label('Pearson Correlation', fontsize=10)
    ax.set_title('Return Correlation Matrix', fontsize=14)
    fig.tight_layout()
    _save_fig('fig4_correlation_heatmap.png')

def figure_5_garch_cond_vol(models: dict, log_ret: pd.DataFrame):
    log.info('Plotting Figure 5: GARCH conditional volatility ...')
    selected = ['ETERNAL.NS', 'PAYTM.NS', 'HINDUNILVR.NS', 'BRITANNIA.NS']
    available = [s for s in selected if s in models]
    fig, ax = plt.subplots(figsize=FIG_SIZE)
    colors_sel = ['#E63946', '#F4A261', '#2A9D8F', '#264653']
    for col, color in zip(available, colors_sel):
        res = models[col]
        cond_vol = res.conditional_volatility / 100
        ax.plot(cond_vol.index, cond_vol.values, color=color, lw=1.0, label=_short(col))
    ax.set_title('GARCH(1,1) Conditional Volatility', fontsize=14)
    ax.set_ylabel('Conditional Volatility (daily, decimal)')
    ax.legend(fontsize=10, framealpha=0.7)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    _save_fig('fig5_garch_conditional_volatility.png')

def figure_6_structural_breaks(log_ret: pd.DataFrame, breaks_dict: Dict[str, List[str]], example_stock: str = 'ETERNAL.NS'):
    log.info('Plotting Figure 6: Structural breaks for %s ...', _short(example_stock))
    if example_stock not in log_ret.columns:
        example_stock = list(log_ret.columns)[0]
    s = log_ret[example_stock].dropna() * 100
    lbl = _short(example_stock)
    break_dates = breaks_dict.get(lbl, [])
    fig, ax = plt.subplots(figsize=FIG_SIZE)
    ax.plot(s.index, s.values, color='#457B9D', lw=0.8, label='Daily Return (%)')
    for bd in break_dates:
        dt = pd.Timestamp(bd)
        ax.axvline(dt, color='#E63946', lw=1.2, alpha=0.7, linestyle='--')
    ax.set_title(f'Structural Breaks — {lbl} (PELT, pen={PELT_PENALTY:.0f})', fontsize=14)
    ax.set_ylabel('Daily Return (%)')
    ax.legend(fontsize=10, framealpha=0.7)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    _save_fig('fig6_structural_breaks.png')

def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ''
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

def create_excel(desc_stats: pd.DataFrame, stats_tests: pd.DataFrame, garch_results: pd.DataFrame, struct_breaks: pd.DataFrame) -> None:
    log.info('Writing results.xlsx ...')
    with pd.ExcelWriter(str(EXCEL_PATH), engine='openpyxl') as writer:
        desc_stats.to_excel(writer, sheet_name='Descriptive Statistics')
        stats_tests.to_excel(writer, sheet_name='Statistical Tests', index=False)
        garch_results.to_excel(writer, sheet_name='GARCH Results', index=False)
        struct_breaks.to_excel(writer, sheet_name='Structural Breaks')
    from openpyxl import load_workbook
    wb = load_workbook(str(EXCEL_PATH))
    for name in wb.sheetnames:
        _auto_width(wb[name])
    wb.save(str(EXCEL_PATH))
    log.info('Excel file saved: %s', EXCEL_PATH)

def main():
    log.info('=' * 60)
    log.info('Invisible Marketing and Market Volatility')
    log.info('=' * 60)
    prices = download_data(ALL_TICKERS)
    prices = align_and_clean(prices)
    prices = filter_by_obs(prices, 'Raw download')
    log_ret, pct_ret, rolling_vol, ann_vol = compute_returns(prices)
    log.info('')
    log.info('--- DESCRIPTIVE STATISTICS ---')
    desc_stats = calc_descriptive_stats(log_ret, ann_vol)
    log.info('\n%s', desc_stats.round(4).to_string())
    log.info('')
    log.info('--- STATISTICAL TESTS ---')
    stats_tests = build_stats_table(log_ret)
    log.info('\n%s', stats_tests.round(4).to_string())
    log.info('')
    log.info('--- GARCH(1,1) ---')
    garch_results, garch_models = fit_garch11(log_ret)
    log.info('\n%s', garch_results.round(4).to_string())
    log.info('')
    log.info('--- GROUP COMPARISON ---')
    t_stat, t_p, l_stat, l_p = group_volatility_test(ann_vol)
    log.info('Algorithm mean ann. vol: %.2f%%', ann_vol[[t for t in ALGORITHM_TICKERS if t in ann_vol.index]].mean())
    log.info('Traditional mean ann. vol: %.2f%%', ann_vol[[t for t in TRADITIONAL_TICKERS if t in ann_vol.index]].mean())
    log.info('Welch t-test: t = %.4f, p = %.6f', t_stat, t_p)
    log.info('Levene test:  stat = %.4f, p = %.6f', l_stat, l_p)
    log.info('')
    log.info('--- STRUCTURAL BREAKS ---')
    struct_breaks_df, breaks_dict = detect_breaks(log_ret, penalty=PELT_PENALTY)
    log.info('')
    log.info('--- GENERATING FIGURES ---')
    figure_1_prices(prices)
    figure_2_rolling_vol(rolling_vol)
    figure_3_boxplot(log_ret)
    figure_4_corr_heatmap(log_ret)
    figure_5_garch_cond_vol(garch_models, log_ret)
    figure_6_structural_breaks(log_ret, breaks_dict)
    log.info('')
    log.info('--- EXPORTING EXCEL ---')
    create_excel(desc_stats, stats_tests, garch_results, struct_breaks_df)
    log.info('')
    log.info('=== DONE ===')
    log.info('Figures saved in: %s/', FIGURE_DIR)
    log.info('Excel saved as:   %s', EXCEL_PATH)

if __name__ == '__main__':
    main()
